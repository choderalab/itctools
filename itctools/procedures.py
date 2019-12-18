from distutils.version import StrictVersion  # For version testing
from datetime import datetime
import re
import numpy
import openpyxl  # Excel spreadsheet I/O (for Auto iTC-200)
from openpyxl import Workbook
import logging

from .itctools import ureg, Quantity, compute_rm
from .labware import PipettingLocation


class ITCProtocol(object):

    def __init__(self, name, sample_prep_method, itc_method, analysis_method, experimental_conditions, injections):
        """
        Parameters
        ----------
        name : str
           The name of the protocol.
        sample_prep_method : str
           The name of the 'SamplePrepMethod' to be written to the Excel file for the Auto iTC-200.
        itc_method : str
           The name of the 'ItcMethod' to be written to the Excel file for the Auto iTC-200.
        analysis_method : str
           The name of the 'AnalysisMethod' to be written to the Excel file for the Auto iTC-200.
        experimental_conditions : dict
            Dictionary containing the experimental conditions
                target_temperature : int in degrees celsius
                equilibration_time : int in seconds
                stir_rate : int in rpm
                reference_power: int in microcalories / second
        injections : list of dict
            list of injections
            With each injection : dict
            volume_inj : float  in microliters
                total volume to be injected
            duration_inj : float in seconds
                duration of the injection phase
            spacing : int in seconds
                time to the next injection
            filter_period : float
                typically 0.5


        """
        # Input validation by type assertion
        assert isinstance(name, str)
        assert isinstance(sample_prep_method, str)
        assert isinstance(itc_method, str)
        assert isinstance(analysis_method, str)
        assert isinstance(experimental_conditions, dict)
        for exp_con in experimental_conditions.values():
            assert isinstance(exp_con, int)
        assert isinstance(injections, list)
        for inj in injections:
            assert isinstance(inj, dict)

        # Strip extension from string if added by user
        self.name = re.sub('\.inj$', '', name)
        self.sample_prep_method = sample_prep_method
        self.itc_method = itc_method
        self.analysis_method = analysis_method
        self.experimental_conditions = experimental_conditions
        self.injections = injections
        self.experimental_conditions['num_inj'] = len(injections)

    def export_inj_file(self):
        """Export the ITC protocol as an origin .inj protocol file

        Note:
        The file format supports some extra options, which currently aren't supported in this library.
        """
        import textwrap

        header=textwrap.dedent("""\
        ITC
        {num_inj}
        NOT
        {target_temperature}
        {equilibration_time}
        {stir_rate}
        {reference_power}
        2
        False,True,True
        """)

        injection_line = "{volume_inj},{duration_inj},{spacing},{filter_period}\n"

        with open(self.name + '.inj', 'w') as inj_file:
            # Fill in the details
            inj_file.write(header.format(**self.experimental_conditions))

            for injection in self.injections:
                inj_file.write(injection_line.format(**injection))


# In case these will diverge at some point, declare alias for Mixture usage.
HeatOfMixingProtocol = ITCProtocol


class ITCExperiment(object):
    """ A single ITC experiment, containing all experimental parameters.
    """
    def __init__(
            self,
            name,
            syringe_source,
            cell_source,
            protocol,
            cell_volume,
            buffer_source=None,
            syringe_concentration=None,
            cell_concentration=None,
            ):
        """
        Parameters
        ----------
        name : str
           Name of the ITC experiment.
        syringe_source : Solution
           Source for syringe solution.
        cell_source : Solution
           Source for cell solution.
        protocol : ITCProtocol
           Protocol to be used for ITC experiment and analysis.
        cell_volume : float
           The volume of the sample cell in microliters
        buffer_source : Labware
           Source for buffer.
        syringe_concentration : pint Quantity with units compatible with moles/liter, optional, default=None
           If specified, syringe source will be diluted to specified concentration.
           Buffer source must be specified.
        cell_concentration : pint Quantity with units compatible with moles/liter, optional, default=None
           If specified, cell source will be diluted to specified concentration.
           Buffer source must be specified.

        WARNING
        -------
        Do not change class member fields after initialization.  Dilution factors are not recomputed.

        """
        self.name = name
        self.syringe_source = syringe_source
        self.cell_source = cell_source
        self.protocol = protocol
        self.cell_volume = cell_volume

        # Store data.
        self.buffer_source = buffer_source
        self.syringe_concentration = syringe_concentration
        self.cell_concentration = cell_concentration

        # Compute dilution factors.
        self.syringe_dilution_factor = None
        self.cell_dilution_factor = None

        if syringe_concentration is not None:
            self.syringe_dilution_factor = syringe_concentration / syringe_source.concentration
            self.syringe_concentration = syringe_concentration

        if cell_concentration is not None:
            self.cell_dilution_factor = cell_concentration / cell_source.concentration
            self.cell_concentration = cell_concentration

        # If dilution is required, make sure buffer source is specified.
        if self.syringe_dilution_factor is not None:
            if buffer_source is None:
                raise Exception(
                    "buffer must be specified if either syringe or cell concentrations are specified")
            if self.syringe_dilution_factor > 1.0:
                raise Exception(
                    "Requested syringe concentration (%s) is greater than syringe source concentration (%s)." %
                    (str(syringe_concentration), str(
                        syringe_source.concentration)))

        if self.cell_dilution_factor is not None:
            if buffer_source is None:
                raise Exception(
                    "buffer must be specified if either syringe or cell concentrations are specified")
            if self.cell_dilution_factor > 1.0:
                raise Exception(
                    "Requested cell concentration (%s) is greater than cell source concentration (%s)." %
                    (str(cell_concentration), str(
                        cell_source.concentration)))

    @staticmethod
    @ureg.wraps(ret=ureg.millimole / ureg.liter, args=[ureg.liter/ureg.millimole, ureg.millimole/ureg.liter, ureg.millimole/ureg.liter, None])
    def _complex_concentration(Ka, total_M, total_L, bound_guess=0.5):
        """
        Compute the equilibrium concentrations of each complex species for N ligands competitively binding to a receptor.
        ARGUMENTS
        Ka Quantity - Ka is the association constant for receptor and ligand species n (1/M)
        M - the concentration of the macromolecule
        L - the concentration of ligand
        bound_guess - guess for the initial bound fraction

        Returns
        concentration of the complex
        """
        from scipy.optimize import minimize_scalar

        def _guess_concentration(complex_estimate, total_M, total_L, Ka):
            Ka_estimate = ((complex_estimate) / ((total_M - complex_estimate) * (total_L - complex_estimate)))
            return abs(Ka - Ka_estimate)

        bounds = tuple([0, min([total_L, total_M])])

        result = minimize_scalar(_guess_concentration, bounds=bounds, args=tuple([total_M, total_L, Ka]), method='Bounded', options=dict(disp=True))

        if not result.success:
            raise RuntimeError("Failed to converge equilibrium concentration.")

        return result.x

    def simulate(self, Ka, plot=True, plot_complex=True, macromol_titrant=False, logscale=False, filename=''):
        """Perform a simulation of the experiment"""

        ninj = self.protocol.experimental_conditions['num_inj']
        molar_ratios = Quantity(numpy.zeros(ninj), 'dimensionless')
        ligand_ratios = Quantity(numpy.zeros(ninj), 'dimensionless')
        macro_ratios = Quantity(numpy.zeros(ninj), 'dimensionless')
        complex_concentrations = Quantity(numpy.zeros(ninj), 'millimole / liter')

        macromolecule_concentrations = Quantity(numpy.zeros(ninj), 'millimole / liter')
        titrant_concentration = Quantity('0.0 mole / liter')
        titrand_concentration = self.cell_concentration


        # Calculate the new concentrations after each injection and store the ratios
        for index, injection in enumerate(self.protocol.injections):

            titrand_amount = titrand_concentration * self.cell_volume
            titrant_amount = titrant_concentration * self.cell_volume + self.syringe_concentration * injection['volume_inj']
            new_volume = self.cell_volume + injection['volume_inj']

            # instantaneous mixing assumed, part of both wasted with each injection
            titrant_concentration = titrant_amount / new_volume
            titrand_concentration = titrand_amount / new_volume

            # If the macromolecule is actually the titrant or not.
            if macromol_titrant:
                macromol_conc = titrant_concentration
                lig_conc = titrand_concentration
            else:
                macromol_conc = titrand_concentration
                lig_conc = titrant_concentration

            ratio = macromol_conc / lig_conc

            molar_ratios[index]= ratio
            complex_conc = self._complex_concentration(Ka,macromol_conc, lig_conc)
            complex_concentrations[index] = complex_conc
            macromolecule_concentrations[index] = macromol_conc - complex_conc
            ligand_ratios[index], macro_ratios[index] = [complex_conc/ (lig_conc - complex_conc) , complex_conc / (macromol_conc - complex_conc)]

        if plot and plot_complex:
            self._plot_simulation(molar_ratios, list(zip(ligand_ratios, macro_ratios)), logscale=logscale, filename=filename)
        elif plot:
            self._plot_simulation(molar_ratios, filename=filename)

        return molar_ratios



    def _plot_simulation(self, molar_ratios, complex_ratios=None, logscale=False, filename=''):
        """Plot the heats from a simulated experiment"""

        import matplotlib.pyplot as plt

        # Only import seaborn if available
        try:
            import seaborn
        except ImportError:
            pass


        fig = plt.figure()
        axtotal = plt.subplot(111)

        graphs = list()
        axcomplex = axtotal.twinx()
        graphs.append(axtotal.plot(range(len(molar_ratios)), [ratio.to('dimensionless') for ratio in molar_ratios],
                                   label='Total Macromolecule/Ligand ratio', c='dimgray'))

        axtotal.set_ylabel('Ratio of Totals')

        if complex_ratios is not None:
            ligand_ratios, macromol_ratios = list(zip(*complex_ratios))

            #Plot complex / free ligand ratio per inj
            graphs.append(axcomplex.plot(range(len(complex_ratios)),
                          [ratio.to('dimensionless') for ratio in ligand_ratios],
                          label='Complex/Ligand ratio', c='crimson'))

            # Plot complex / free macromolecule ratio per inj
            graphs.append(axcomplex.plot(range(len(complex_ratios)),
                          [ratio.to('dimensionless') for ratio in macromol_ratios],
                          label='Complex/Macromolecule ratio', c='lightskyblue'))

            axcomplex.set_xlabel('Injection')
            axcomplex.set_ylabel('Complex/Free ratio')
            if logscale:
                axcomplex.set_yscale('log')
        # flatten graphs
        graphs = [item for sublist in graphs for item in sublist]
        plotlabels = [g.get_label() for g in graphs]

        axtotal.legend(graphs,plotlabels, loc=0)

        if filename:
            plt.savefig(filename, dpi=300)
        else:
            plt.show()

        plt.close()


class ITCHeuristicExperiment(ITCExperiment):
    """
    A single ITC experiment, containing all experimental parameters.

    Implements the heuristic equation from

    Optimizing Experimental Parameters in Isothermal Titration Calorimetry
    Joel Tellinghuisen
    The Journal of Physical Chemistry B 2005 109 (42), 20027-20035

    """

    def heuristic_syringe(self, association_constant, throw_away=1, approx=False, strict=True):
        """
        Optimize syringe concentration using heuristic equation.

        Parameters
        ----------
        association_constant : pint Quantity with units compatible with liters/moles
            Association constant of titrant from titrand
        throw_away: int
           Number of injections to consider as throw_away injections
        approx: bool
            Use approximate equation [X]_s = R_m * [M]0 V/(m*v) if True
            else, use exact equation [X]_s = R_m * [M]_0 (1- exp(-mv/V0))^-1

        Reference
        ---------
        http://dx.doi.org/10.1016/j.ab.2011.03.024
        """
        assert isinstance(throw_away, int)
        assert throw_away >= 0

        m = self.protocol.experimental_conditions['num_inj']
        if throw_away:
            injections = self.protocol.injections[throw_away:]
        else:
            injections = self.protocol.injections

        v = sum(injection['volume_inj'] for injection in injections) / len(injections)
        v0 = self.cell_volume

        # c = [M]_0 * Ka
        c = self.cell_concentration * association_constant

        # R_m = 6.4/c^0.2 + 13/c
        rm = compute_rm(c)

        if strict and rm < 1.1:
            raise ValueError("Value of Rm should be greater than 1: %s" % rm)
        elif (not strict) and rm < 1.1:
            logging.warning("High affinity or high cell concentration: Rm = %f. Using 1.1." % rm)
            rm = 1.1

        if not approx:
            # Use exact equation [X]_s = R_m * [M]_0 (1- exp(-mv/V0))^-1
            self.syringe_concentration = rm * self.cell_concentration * \
                numpy.power(1 - (numpy.exp(-1 * m * v / v0)), -1)
        else:
            # Use approximate equation [X]_s = R_m * [M]0 V/(m*v)
            self.syringe_concentration = rm * \
                self.cell_concentration * v0 / (m * v)

        # compute the dilution factors
        self.syringe_dilution_factor = numpy.float(
            self.syringe_concentration /
            self.syringe_source.concentration)

        return rm

    def rescale(self, sfactor=None, cfactor=None, tfactor=None):
        """Rescale the concentrations while keeping same ratios, adjust in case they are larger than the source.

        Parameters
        ----------
        sfactor : float
            if not None, also scale syringe concentrations by this factor.
        cfactor : float
            if not None, also scale cell concentration by this factor.
        tfactor : float
            if not None, scale all concentrations by this factor.

        Returns
        -------
        float
            the final factor by which everything was scaled
        """

        # if syringe concentration is larger than stock
        if sfactor is None:
            sfactor = 1.00
        elif self.syringe_concentration is None:
            raise RuntimeWarning("Attempted to rescale nonexistent solution.")
        if cfactor is None:
            cfactor = 1.00
        if tfactor is None:
            tfactor = 1.00

        # If there is no syringe concentration, don't attempt to scale.
        if self.syringe_concentration is not None:

            # Syringe concentration scaling factor
            sfactor *= tfactor
            if self.syringe_concentration > self.syringe_source.concentration:
                # Multiply original factor by the necessary rescaling
                sfactor *= self.syringe_source.concentration / self.syringe_concentration
            # scale down to stock
            sfactor *= tfactor
            self.syringe_concentration *= sfactor
            # cell is scaled by same factor
            self.cell_concentration *= sfactor

        # Cell concentration scaling factor
        cfactor *= tfactor
        try:
            if self.cell_concentration > self.cell_source.concentration:
                # Multiply original factor by the necessary rescaling
                cfactor *= self.cell_source.concentration / self.cell_concentration
            # scale down to stock
            self.cell_concentration *= cfactor
            # syringe is scaled by same factor
            if self.syringe_concentration is not None:
                self.syringe_concentration *= cfactor
            # recompute dilution factor
            self.cell_dilution_factor = self.cell_concentration / self.cell_source.concentration

        except AttributeError as err:
            # Labware has no concentration (buffer)
            print("WARNING, cell cannot be rescaled. This may still be desired if cell is not buffer or water.")
            print("Full error details: \n %s" % err)

        # recompute dilution factor
        if self.syringe_concentration is not None:
            self.syringe_dilution_factor = self.syringe_concentration / self.syringe_source.concentration

        return sfactor * cfactor


class HeatOfMixingExperiment(object):

    def __init__(self, name, cell_mixture, syringe_mixture, protocol):
        """
        Parameters
        ----------
        name : str
           Name of the ITC experiment.
        cell_mixture : SimpleMixture
            Initial composition of the cell mixture
        syringe_mixture : SimpleMixture
            Composition of the syringe mixture
        protocol : ITCProtocol
           Protocol to be used for ITC experiment and analysis.
        """
        self.name = name
        self.cell_mixture = cell_mixture
        self.syringe_mixture = syringe_mixture
        self.protocol = protocol


class ITCExperimentSet(object):

    def __init__(self, name):
        """
        Parameters
        ----------
        name : str
           Name of the experiment set.

        """
        self.name = name  # name of the experiment
        self.experiments = list()  # list of experiments to set up
        # ITC plates available for use in experiment
        self.destination_plates = list()

        self._validated = False

    def addDestinationPlate(self, plate):
        """
        Add the specified destination plate to the plate set usable for setting up ITC experiments.

        Parameters
        ----------
        plate : Labware
           The empty ITC destination plate to add to the experiment set.

        """

        # TODO: Check if specified plate is allowed type of labware for use in
        # Auto iTC-200.
        self.destination_plates.append(plate)

    def addExperiment(self, experiment):
        """
        Add the specified ITC experiment to the experiment set.

        Parameters
        ----------
        experiment : ITCExperiment
           The ITC experiment to add.

        """
        self.experiments.append(experiment)

    def _wellIndexToName(self, index):
        """
        Return the 96-well name (e.g. 'A6', 'B7') corresponding to Tecan well index.

        Parameters
        ----------
        index : int
           Tecan well index (back to front, left to right), numbered from 1..96

        Returns
        -------
        well_name : str
           Well name for ITC plate (e.g. 'A6'), numbered from A1 to H12

        """
        row = int((index - 1) % 8)
        column = int((index - 1) / 8)
        rownames = 'ABCDEFGH'
        well_name = rownames[row] + '%d' % (column + 1)
        return well_name

    class ITCData(object):

        def __init__(self):
            fieldnames = [
                'DataFile',
                'SampleName',
                'SamplePrepMethod',
                'ItcMethod',
                'AnalysisMethod',
                'CellConcentration',
                'PipetteConcentration',
                'CellSource',
                'PipetteSource',
                'PreRinseSource',
                'SaveSampleDestination']
            for fieldname in fieldnames:
                setattr(self, fieldname, None)

    class TecanData(object):

        def __init__(self):
            fieldnames = [
                'cell_destination',
                'cell_platename',
                'cell_wellindex',
                'syringe_plateindex',
                'syringe_platename',
                'syringe_wellindex']
            for fieldname in fieldnames:
                setattr(self, fieldname, None)

    def _resetTrackedQuantities(self):
        self._tracked_quantities = dict()

    def _trackQuantities(self, thing, volume):
        try:
            name = thing.name
        except:
            name = thing.RackLabel

        if name in self._tracked_quantities:
            self._tracked_quantities[name] += volume
        else:
            self._tracked_quantities[name] = volume

    def _allocate_destinations(self):
        # Make a list of all the possible destination pipetting locations.
        # TODO: Change this to go left-to-right in ITC plates?
        self.destination_locations = list()
        for (plate_index, plate) in enumerate(self.destination_plates):
            PlateNumber = plate_index + 1
            for index in range(96):
                Position = index + 1
                WellName = self._wellIndexToName(Position)
                location = PipettingLocation(
                    plate.RackLabel,
                    plate.RackType,
                    Position)
                # Add plate number and well name for Auto iTC-200.
                location.PlateNumber = PlateNumber
                location.WellName = WellName
                self.destination_locations.append(location)

    @staticmethod
    def human_readable(source):
        """
        Generate a human-readable position name

        Parameters
        ----------
        source : obj
            Object with .RackLabel and .RackType fields

        Returns
        -------
        description : str
            The human-readable description
            'SourcePlate well A1'
            'Buffer trough'
        """
        positionless = False
        if not hasattr(source, 'Position'):
            positionless = True

        if source.RackType == '5x3 Vial Holder':
            ny = 3; nx = 5
        elif source.RackType == 'Trough 100ml':
            ny = 8; nx = 1
        elif source.RackType == 'ITC Plate':
            ny = 8; nx = 12
        else:
            raise Exception(f"Don't know racktype '{racktype}'")

        description = ''
        if positionless:
            description = f'{source.RackLabel} trough'
        else:
            # TODO: Check that position is not outside range
            wells = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
            y = int((source.Position-1)/ny)
            x = source.Position-y*ny
            wellname = f'{wells[y]}{x}'
            description = f'{source.RackLabel} well {wellname}'

        return description

    def validate(self, print_volumes=False, omit_zeroes=True, vlimit=10.0, human_readable_log=None):
        """
        Validate that the specified set of ITC experiments can actually be set up, raising an exception if not.

        Additional experiment data fields (tecandata, itcdata)

        This method splits pipetting operations into LiHa (buffer, receptor) and aLiHa (ligand).
        The ``self.worklists[key]`` attribute contains the 'LiHa' and 'aLiHa' worklists indexed by 'key'

        Parameters
        ----------
        volumes : bool
            Print out pipetting volumes (default=True)
        omit_zeroes : bool
            Omit operations with volumes below vlimit (default = True)
        vlimit : float
            Minimal volume for pipetting operation in microliters (default = 10.0)
        human_readable_log : file-like object, optional, default=None
            If specified, will write a human-readable pipetting log to the specified file-like object
        """

        # TODO: Try to set up experiment, throwing exception upon failure.

        # Make a list of all the possible destination pipetting locations.
        self._allocate_destinations()

        def transfer(source, dest, volume):
            """Write human-readable quanties if desired."""
            operation = f'* Transfer {volume:.1f} uL from {self.human_readable(source)} to {self.human_readable(dest)}\n'
            if human_readable_log:
                # TODO: Streamline this with a single transfer command
                human_readable_log.write(operation)
            return operation

        # Build worklist scripts
        self.worklists = dict()
        self.worklists['LiHa'] = ""
        self.worklists['aLiHa'] = ""
        aLiHa_max_volume = 50 # 50 uL tips

        # Reset tracked quantities.
        self._resetTrackedQuantities()

        for (experiment_number, experiment) in enumerate(self.experiments):
            # volume logging
            volume_report = str()

            experiment.experiment_number = experiment_number

            # volume logging
            volume_report += f"Experiment {experiment.experiment_number+1} : {experiment.name} : \n"

            itcdata = ITCExperimentSet.ITCData()
            tecandata = ITCExperimentSet.TecanData()

            #
            # Cell contents
            #

            # Find a place to put cell contents.
            if len(self.destination_locations) == 0:
                raise Exception(
                    "Ran out of destination plates for experiment %d / %d" %
                    (experiment_number, len(self.experiments)))
            tecandata.cell_destination = self.destination_locations.pop(0)

            cell_volume = 400.0  # microliters # TODO: Let this be a class field; manual recommends 370 uL
            transfer_volume = cell_volume

            if experiment.cell_dilution_factor is not None:
                # Compute buffer volume needed.
                buffer_volume = cell_volume * (
                    1.0 - experiment.cell_dilution_factor)
                transfer_volume = cell_volume - buffer_volume

                # Schedule buffer transfer.
                tipmask = 1

                if (buffer_volume > 0.01 or not omit_zeroes):
                    # TODO: Remove duplicated code by rolling this into a method
                    if (buffer_volume < aLiHa_max_volume):
                        # Use aLiHa for better precision
                        pipette = 'aLiHa'
                    else:
                        pipette = 'LiHa'

                    source = experiment.buffer_source
                    dest = tecandata.cell_destination

                    self.worklists[pipette] += 'A;%s;;%s;%d;;%f;;;%d\r\n' % (
                        source.RackLabel, source.RackType, 1,
                        buffer_volume, tipmask)
                    self.worklists[pipette] += 'D;%s;;%s;%d;;%f;;;%d\r\n' % (
                        tecandata.cell_destination.RackLabel, dest.RackType, dest.Position,
                        buffer_volume, tipmask)

                    volume_report += transfer(source, dest, buffer_volume)

                    # no wash if no actions taken
                    self.worklists[pipette] += 'W;\r\n'  # queue wash tips
                    self._trackQuantities(
                        experiment.buffer_source,
                        buffer_volume *
                        ureg.microliters)

            # Schedule cell solution transfer.
            tipmask = 2
            try:
                # Assume source is Solution.
                if (transfer_volume > 0.01 or not omit_zeroes):
                    if (transfer_volume < aLiHa_max_volume):
                        # Use aLiHa for better precision
                        pipette = 'aLiHa'
                    else:
                        pipette = 'LiHa'
                    source = experiment.cell_source.location
                    self.worklists[pipette] += 'A;%s;;%s;%d;;%f;;;%d\r\n' % (
                            source.RackLabel, source.RackType, source.Position, transfer_volume, tipmask)
            except:
                # Assume source is Labware.
                if (transfer_volume > 0.01 or not omit_zeroes):
                    if (transfer_volume < aLiHa_max_volume):
                        # Use aLiHa for better precision
                        pipette = 'aLiHa'
                    else:
                        pipette = 'LiHa'
                    source = experiment.cell_source
                    self.worklists[pipette] += 'A;%s;;%s;%d;;%f;;;%d\r\n' % (
                        source.RackLabel, source.RackType, 2, transfer_volume, tipmask)

            if (transfer_volume > 0.01 or not omit_zeroes):
                dest = tecandata.cell_destination
                self.worklists[pipette] += 'D;%s;;%s;%d;;%f;;;%d\r\n' % (
                    dest.RackLabel, dest.RackType, dest.Position, transfer_volume, tipmask)

                volume_report += transfer(source, dest, transfer_volume)

                # no wash if no actions taken
                if pipette != None:
                    self.worklists[pipette] += 'W;\r\n'  # wash tips / change DiTi
                self._trackQuantities(
                    experiment.cell_source,
                    transfer_volume *
                    ureg.microliters)

            #
            # Syringe contents
            #

            # Find a place to put syringe contents.
            if len(self.destination_locations) == 0:
                raise Exception(
                    "Ran out of destination plates for experiment %d / %d" %
                    (experiment_number, len(self.experiments)))
            tecandata.syringe_destination = self.destination_locations.pop(0)

            syringe_volume = 150.0  # microliters (120 uL minimimum)
            transfer_volume = syringe_volume

            if experiment.syringe_dilution_factor is not None:
                if (experiment.syringe_dilution_factor > 1.0) or (experiment.syringe_dilution_factor < 0.0):
                    raise Exception(f'experiment.syringe_dilution_factor = {experiment.syringe_dilution_factor}')

                # Compute buffer volume needed.
                buffer_volume = syringe_volume * (
                    1.0 - experiment.syringe_dilution_factor)
                transfer_volume = syringe_volume - buffer_volume

                # Schedule buffer transfer.
                tipmask = 4
                pipette = None
                if (buffer_volume > 0.01 or not omit_zeroes):

                    if (buffer_volume < aLiHa_max_volume):
                        # Use aLiHa for better precision
                        pipette = 'aLiHa'
                    else:
                        pipette = 'LiHa'

                    source = experiment.buffer_source
                    dest = tecandata.syringe_destination
                    self.worklists[pipette] += 'A;%s;;%s;%d;;%f;;;%d\r\n' % (
                        source.RackLabel, source.RackType, 3, buffer_volume, tipmask)
                    self.worklists[pipette] += 'D;%s;;%s;%d;;%f;;;%d\r\n' % (
                        dest.RackLabel, dest.RackType, dest.Position, buffer_volume, tipmask)

                    volume_report += transfer(source, dest, buffer_volume)

                    # no wash if no actions taken
                    if pipette != None:
                        self.worklists[pipette] += 'W;\r\n'  # wash tips / change DiTi
                    self._trackQuantities(
                        experiment.buffer_source,
                        buffer_volume *
                        ureg.microliters)

            # Schedule syringe solution transfer.
            tipmask = 8
            pipette = None
            try:
                # Assume source is Solution: use aLiHa
                if (transfer_volume > 0.01 or not omit_zeroes):
                    if (transfer_volume < aLiHa_max_volume):
                        # Use aLiHa for better precision
                        pipette = 'aLiHa'
                    else:
                        pipette = 'LiHa'
                    source = experiment.syringe_source.location
                    self.worklists[pipette] += 'A;%s;;%s;%d;;%f;;;%d\r\n' % (
                        source.RackLabel,
                        source.RackType,
                        source.Position,
                        transfer_volume, tipmask)
            except:
                if (transfer_volume < aLiHa_max_volume):
                    # Use aLiHa for better precision
                    pipette = 'aLiHa'
                else:
                    pipette = 'LiHa'
                # Assume source is Labware
                source = experiment.syringe_source
                if transfer_volume > 0.01 or not omit_zeroes:
                    self.worklists[pipette] += 'A;%s;;%s;%d;;%f;;;%d\r\n' % (
                        source.RackLabel, source.RackType, 4, transfer_volume, tipmask)

            if (transfer_volume > 0.01 or not omit_zeroes):
                dest = tecandata.syringe_destination
                self.worklists[pipette] += 'D;%s;;%s;%d;;%f;;;%d\r\n' % (
                    dest.RackLabel, dest.RackType, dest.Position, transfer_volume, tipmask)
                self.worklists[pipette] += 'W;\r\n'  # wash tips / change DiTi
                self._trackQuantities(
                    experiment.syringe_source,
                    transfer_volume *
                    ureg.microliters)

            volume_report += transfer(source, dest, transfer_volume)

            # Finish worklist section.
            #pipette = 'LiHa'
            #self.worklists[pipette] += 'B;\r\n'  # execute queued batch of commands

            # Create datafile name.
            now = datetime.now()
            datecode = now.strftime("%Y%m%d")
            seriescode = 'a'  # TODO: Use intelligent coding?
            indexcode = '%d' % (experiment_number + 1)
            itcdata.DataFile = datecode + seriescode + indexcode + '.itc'

            itcdata.SampleName = experiment.name
            itcdata.SamplePrepMethod = experiment.protocol.sample_prep_method
            itcdata.ItcMethod = experiment.protocol.itc_method
            itcdata.AnalysisMethod = experiment.protocol.analysis_method

            try:
                itcdata.CellConcentration = experiment.cell_concentration / (ureg.millimole / ureg.liter)
            except:
                itcdata.CellConcentration = 0

            try:
                itcdata.PipetteConcentration = experiment.syringe_concentration / (ureg.millimole / ureg.liter)
            except:
                itcdata.PipetteConcentration = 0

            itcdata.CellSource = 'Plate%d, %s' % (
                tecandata.cell_destination.PlateNumber, tecandata.cell_destination.WellName)
            itcdata.PipetteSource = 'Plate%d, %s' % (
                tecandata.syringe_destination.PlateNumber, tecandata.syringe_destination.WellName)

            # TODO: Autodetect if prerinse is used.
            itcdata.PreRinseSource = ''

            # TODO: Autodetect if sample destination is used.
            itcdata.SaveSampleDestination = itcdata.CellSource

            # Store Tecan and Excel data for this experiment.
            experiment.tecandata = tecandata
            experiment.itcdata = itcdata
            if print_volumes:
                print(volume_report)

        # Report tracked quantities.
        print("Necessary volumes:")
        keys = sorted(self._tracked_quantities.keys())
        for key in keys:
            print("%32s %12.3f mL" % (key, self._tracked_quantities[key] / ureg.milliliters))

        # Report expected waste
        print("Expected waste (3% of total):")
        keys = sorted(self._tracked_quantities.keys())
        for key in keys:
            print("%32s %12.3f mL" % (key, 0.03 * self._tracked_quantities[key] / ureg.milliliters))

        # Set validated flag.
        self._validated = True

    def writeTecanWorklist(self, prefix):
        """
        Write the Tecan worklist for the specified experiment set.

        Parameters
        ----------
        prefix : str
           The prefix of the Tecan worklist files to write.

        """
        self.validate()
        for key in self.worklists.keys():
            with open('%s-%s.gwl' % (prefix, key) , 'w') as outfile:
                outfile.write(self.worklists[key])

    def writeAutoITCExcel(self, filename):
        """
        Write the Excel file for the specified experiment set to be loaded into the Auto iTC-200.

        Parameters
        ----------
        filename : str
           The name of the Excel file to write.

        """

        if not self._validated:
            self.validate()

        # Create new Excel spreadsheet.

        wb = Workbook()

        # Create plate sheet.
        ws = wb.get_active_sheet()
        ws.title = 'plate'

        # Openpyxl version incompatibility fix
        if StrictVersion(openpyxl.__version__) >= StrictVersion('2.0.0'):
            row = 1
            start = 1
        else:
            row = 0
            start = 0
        # Create header.
        fieldnames = [
            'DataFile',
            'SampleName',
            'SamplePrepMethod',
            'ItcMethod',
            'AnalysisMethod',
            'CellConcentration',
            'PipetteConcentration',
            'CellSource',
            'PipetteSource',
            'PreRinseSource',
            'SaveSampleDestination']
        for (column, fieldname) in enumerate(fieldnames, start=start):
            ws.cell(row=row, column=column).value = fieldname

        # Create experiments.
        for experiment in self.experiments:
            row += 1
            for (column, fieldname) in enumerate(fieldnames, start=start):
                value = getattr(experiment.itcdata, fieldname)
                try:
                    if value.dimensionless:
                        value = float(value)
                except AttributeError:
                    pass

                ws.cell(
                    row=row,
                    column=column).value = value

        # Write workbook.
        wb.save(filename)


class HeatOfMixingExperimentSet(ITCExperimentSet):

    """
    Set up experiments to calculate the heat of mixing for a mixture.

    TODO: Work out the concepts
    """

    def __init__(self, name):
        """
        Parameters
        ----------
        name : str
            Identifier for the experiment set.

        """
        super(HeatOfMixingExperimentSet, self).__init__(name)
        self._worklist_complete = False
        self._autoitc_complete = False
        self._validated = False

    @staticmethod
    def _worklist_line(operation, tipmask, labware, index, volume):
        """
        Write an aspirate or dispense operation to the worklist.

        :param operation: Type of operation (A, D)
        :type operation: str
        :param tipmask: Dictionary containing the tip masks per solution
        :type tipmask:  dict
        :param labware: Target location that contains RackLabel, RackType and Position
        :type labware: PipettingLocation
        :param index: Index of the mixture component
        :type index: int
        :param volume: volume to be dispensed
        :type volume: pint.Quantity
        :return: The worklist line
        :rtype: str
        """
        assert operation in "AD"

        # Aspire operation, look for the mixture component based on index and extract its location
        if operation == "A":
            location = labware.locations[index]
        # Dispense operation, location is given
        elif operation == "D":
            location = labware

        worklist_line = '%s;%s;;%s;%d;;%f;;;%d\r\n' % (operation,
                                                          location.RackLabel,
                                                          location.RackType,
                                                          location.Position,
                                                          volume[index],
                                                          tipmask)

        return worklist_line

    def populate_worklist(
            self,
            cell_volume=400.0 *
            ureg.microliters,
            syringe_volume=120.0 *
            ureg.microliters):
        """
        Build the worklists for heat of mixing experiments.

        This method only uses the `LiHa`.

        Parameters
        ----------
        cell_volume : pint Quantity with units compatible with microliters
            Total volume to prepare for cell in microliters  (opt., default = 400.0 * microliters )
        syringe_volume : pint Quantity with units compatible with microliters
            Total volume to prepare for syringe in microliters (default = 120.0 * microliters )
        """

        # Make a list of all the possible destination pipetting locations.
        self._allocate_destinations()

        # Build worklist script.
        self.worklists = dict()
        self.worklists['LiHa'] = ""

        # Reset tracked quantities.
        self._resetTrackedQuantities()

        for (experiment_number, experiment) in enumerate(self.experiments):

            # Assign experiment number
            experiment.experiment_number = experiment_number
            tecandata = HeatOfMixingExperimentSet.TecanData()

            # Ensure there are ITC wells available
            if len(self.destination_locations) == 0:
                raise Exception(
                    "Ran out of destination plates for experiment %d / %d" %
                    (experiment_number, len(self.experiments)))
            tecandata.cell_destination = self.destination_locations.pop(0)

            # Calculate volumes per component for cell mixture
            cell_volumes = list()
            for cellfrac in experiment.cell_mixture.volumefractions:
                # Ensure units are correct
                cell_volumes.append(
                    float(
                        cell_volume *
                        cellfrac /
                        ureg.microliters))

            # Calculate volumes per component for syringe mixture
            syr_volumes = list()
            for syrfrac in experiment.syringe_mixture.volumefractions:
                # Ensure units are correct
                syr_volumes.append(
                    float(
                        syringe_volume *
                        syrfrac /
                        ureg.microliters))

            allcomponents = experiment.cell_mixture.components + experiment.syringe_mixture.components
            allcomponents = list(set(allcomponents))

            # dictionary for tip masks
            dictips = dict()
            for val, key in enumerate(allcomponents):
                dictips[key] = 2 ** val

            # Start mixing up our cell volume
            for i in range(len(experiment.cell_mixture.components)):
                self.worklists['LiHa'] += self._worklist_line('A', dictips[experiment.cell_mixture.components[i]], experiment.cell_mixture, i, cell_volumes)

                self.worklists['LiHa'] += self._worklist_line('D', dictips[experiment.cell_mixture.components[i]], tecandata.cell_destination, i, cell_volumes)

                self.worklists['LiHa'] += 'W;\r\n'  # queue wash tips

                self._trackQuantities(
                    experiment.cell_mixture.components[i],
                    cell_volumes[i] *
                    ureg.microliters)

            # Find a place to put syringe contents.
            if len(self.destination_locations) == 0:
                raise Exception(
                    "Ran out of destination plates for experiment %d / %d" %
                    (experiment_number, len(self.experiments)))
            tecandata.syringe_destination = self.destination_locations.pop(0)

            # Start mixing up our syringe volume
            pipette = 'LiHa'
            for i in range(len(experiment.syringe_mixture.components)):
                self.worklists[pipette] += self._worklist_line('A', dictips[experiment.syringe_mixture.components[i]], experiment.syringe_mixture, i, syr_volumes)

                self.worklists[pipette] += self._worklist_line('D', dictips[experiment.syringe_mixture.components[i]], tecandata.syringe_destination, i, syr_volumes)

                self.worklists[pipette] += 'W;\r\n'  # queue wash tips

                self._trackQuantities(
                    experiment.syringe_mixture.components[i],
                    syr_volumes[i] *
                    ureg.microliters)

            # Finish worklist section.
            self.worklists[pipette] += 'B;\r\n'  # execute queued batch of commands

            # Store Tecan data for this experiment
            experiment.tecandata = tecandata

        # Store the completed worklist, containing all experiments
        self._worklist_complete = True

    def populate_autoitc_spreadsheet(self):
        """
        Populate all the fields in the Auto-iTC200 spreadsheet
        """
        # Plate wells need to be assigned in order to generate the spreadsheet
        if not self._worklist_complete:
            raise Exception("Please generate a Tecan Worklist first!")

        for (experiment_number, experiment) in enumerate(self.experiments):

            itcdata = HeatOfMixingExperimentSet.ITCData()
            # Create datafile name.
            now = datetime.now()
            datecode = now.strftime("%Y%m%d")
            seriescode = 'a'  # TODO: Use intelligent coding?
            indexcode = '%d' % (experiment_number + 1)
            prefix = 'CHODERA' # TODO: Allow user to specify prefix
            itcdata.DataFile = prefix + datecode + seriescode + indexcode
            itcdata.SampleName = experiment.name
            itcdata.SamplePrepMethod = experiment.protocol.sample_prep_method
            itcdata.ItcMethod = experiment.protocol.itc_method
            itcdata.AnalysisMethod = experiment.protocol.analysis_method

            # Not a binding experiment, set concentrations to 0
            itcdata.CellConcentration = 0
            itcdata.PipetteConcentration = 0

            itcdata.CellSource = 'Plate%d, %s' % (
                experiment.tecandata.cell_destination.PlateNumber, experiment.
                tecandata.cell_destination.WellName)
            itcdata.PipetteSource = 'Plate%d, %s' % (
                experiment.tecandata.syringe_destination.PlateNumber,
                experiment.tecandata.syringe_destination.WellName)

            # TODO: Autodetect if prerinse is used.
            itcdata.PreRinseSource = ''

            # TODO: Autodetect if sample destination is used.
            itcdata.SaveSampleDestination = itcdata.CellSource

            # Store Excel data for this experiment.
            experiment.itcdata = itcdata
        self._autoitc_complete = True

    def report_quantities(self):
        # Report tracked quantities.
        print("Necessary volumes:")
        keys = sorted(self._tracked_quantities.keys())
        for key in keys:
            print("%32s %12.3f mL" % (key, self._tracked_quantities[key] / ureg.milliliters))

        # Report expected waste
        print("Expected waste (5% of total):")
        keys = sorted(self._tracked_quantities.keys())

        for key in keys:
            print("%32s %12.3f mL" % (key, 0.05 * self._tracked_quantities[key] / ureg.milliliters))

    def validate(self, strict=True):
        """Make sure that necessary steps have been taken before writing to files."""
        if not self._autoitc_complete:
            message = "Auto-iTC200 spreadsheet (.xls) not yet populated!"
            if strict:
                raise RuntimeError(message)
            else:
                print("Warning: %s"% message)
        elif not self._worklist_complete:
            message = "Tecan worklist (.gwl) not yet populated!"
            if strict:
                raise RuntimeError(message)
            else:
                print("Warning: %s"% message)
        else:
            self._validated = True
